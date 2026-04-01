"""
Report generation utilities for Excel and PDF exports with charts
"""
import pandas as pd
from io import BytesIO
import base64
from typing import Dict, List, Any, Optional
import json
from datetime import datetime
import numpy as np

# NOTE: matplotlib and seaborn are heavy plotting libraries. Import them lazily
# inside ReportGenerator methods so the module can be imported and the app can
# start even if plotting libs are not installed. This avoids startup failures
# when report generation is not used.
plt = None
sns = None

try:
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ReportGenerator:
    """Generate reports in various formats with charts"""
    
    def __init__(self):
        self.chart_style = {
            'figure.figsize': (10, 6),
            'axes.titlesize': 14,
            'axes.labelsize': 12,
            'xtick.labelsize': 10,
            'ytick.labelsize': 10,
            'legend.fontsize': 10
        }
        # Lazy import plotting libraries when an instance is created.
        # If imports fail, we set flags but allow the application to continue.
        global plt, sns
        try:
            import matplotlib.pyplot as _plt
            plt = _plt
            plt.rcParams.update(self.chart_style)
        except Exception:
            plt = None

        try:
            import seaborn as _sns
            sns = _sns
            # Only set seaborn palette if seaborn imported and matplotlib present
            if sns is not None and plt is not None:
                sns.set_palette("husl")
        except Exception:
            sns = None
    
    def generate_excel_report(self, data: Dict[str, Any], template_id: str) -> BytesIO:
        """Generate Excel report with charts"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        buffer = BytesIO()
        
        # Create workbook and worksheets
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            self._create_summary_sheet(writer, data, template_id)
            
            # Data sheet
            if 'responses' in data:
                df = pd.DataFrame(data['responses'])
                df.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Charts sheet
            self._create_charts_sheet(writer, data)
        
        buffer.seek(0)
        return buffer
    
    def generate_pdf_report(self, data: Dict[str, Any], template_id: str) -> BytesIO:
        """Generate PDF report with charts"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(f"ICT Assessment Report - {template_id.replace('-', ' ').title()}", title_style))
        story.append(Spacer(1, 12))
        
        # Summary
        if 'summary' in data:
            self._add_summary_to_pdf(story, data['summary'], styles)
        
        # Charts
        chart_images = self._generate_chart_images(data)
        for chart_name, chart_buffer in chart_images.items():
            story.append(Paragraph(f"{chart_name.replace('_', ' ').title()}", styles['Heading2']))
            chart_buffer.seek(0)
            img = RLImage(chart_buffer, width=6*inch, height=4*inch)
            story.append(img)
            story.append(Spacer(1, 12))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, writer, data: Dict[str, Any], template_id: str):
        """Create summary sheet in Excel"""
        summary_data = []
        
        # Report info
        summary_data.append(['Report Type', template_id.replace('-', ' ').title()])
        summary_data.append(['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append(['', ''])
        
        # Key metrics
        if 'summary' in data:
            summary = data['summary']
            summary_data.extend([
                ['Key Metrics', ''],
                ['Total Responses', summary.get('totalResponses', 0)],
                ['Average Infrastructure Score', f"{summary.get('averageInfrastructureScore', 0):.2f}"],
                ['Average Satisfaction', f"{summary.get('averageSatisfaction', 0):.2f}"],
                ['Critical Barriers Count', summary.get('criticalBarriersCount', 0)],
                ['Colleges Covered', summary.get('collegesCount', 0)],
            ])
        
        # Create DataFrame and write to Excel
        df_summary = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format the summary sheet
        workbook = writer.book
        worksheet = writer.sheets['Summary']
        
        # Style headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    def _create_charts_sheet(self, writer, data: Dict[str, Any]):
        """Create charts sheet in Excel"""
        if 'responses' not in data or not data['responses']:
            return
        
        df = pd.DataFrame(data['responses'])
        
        # Create charts sheet
        workbook = writer.book
        charts_sheet = workbook.create_sheet('Charts')
        
        # Generate chart images
        chart_images = self._generate_chart_images(data)
        
        # Add charts to sheet
        row = 1
        for chart_name, chart_buffer in chart_images.items():
            # Add chart title
            charts_sheet.cell(row=row, column=1, value=chart_name.replace('_', ' ').title())
            charts_sheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 2
            
            # Add chart image
            chart_buffer.seek(0)
            img = Image(chart_buffer)
            img.width = 600
            img.height = 400
            charts_sheet.add_image(img, f'A{row}')
            row += 22  # Space for next chart
    
    def _generate_chart_images(self, data: Dict[str, Any]) -> Dict[str, BytesIO]:
        """Generate chart images as BytesIO objects"""
        charts = {}
        
        if 'responses' not in data or not data['responses']:
            return charts
        
        df = pd.DataFrame(data['responses'])

        # Ensure plotting libraries are available
        if plt is None:
            raise ImportError("matplotlib is required for chart generation")
        
        # Infrastructure Distribution Chart
        if 'infrastructure' in df.columns:
            fig, ax = plt.subplots(figsize=(10, 6))
            infra_scores = []
            for resp in data['responses']:
                infra = resp.get('infrastructure', {})
                score = (
                    infra.get('hardwareQuality', 0) +
                    infra.get('softwareAvailability', 0) +
                    infra.get('internetSpeed', 0) +
                    infra.get('digitalCollection', 0)
                ) / 4
                infra_scores.append(score)
            
            plt.hist(infra_scores, bins=10, alpha=0.7, color='skyblue', edgecolor='black')
            plt.title('Infrastructure Score Distribution')
            plt.xlabel('Infrastructure Score')
            plt.ylabel('Frequency')
            plt.grid(True, alpha=0.3)
            
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close()
            charts['infrastructure_distribution'] = buffer
        
        # Satisfaction by College Chart
        if 'collegeName' in df.columns and 'serviceQuality' in df.columns:
            fig, ax = plt.subplots(figsize=(12, 6))
            college_satisfaction = {}
            for resp in data['responses']:
                college = resp.get('collegeName', 'Unknown')
                satisfaction = resp.get('serviceQuality', {}).get('overallSatisfaction', 0)
                if college not in college_satisfaction:
                    college_satisfaction[college] = []
                college_satisfaction[college].append(satisfaction)
            
            colleges = list(college_satisfaction.keys())
            avg_satisfaction = [np.mean(college_satisfaction[c]) for c in colleges]
            
            plt.bar(range(len(colleges)), avg_satisfaction, color='lightcoral')
            plt.title('Average Satisfaction by College')
            plt.xlabel('College')
            plt.ylabel('Average Satisfaction Score')
            plt.xticks(range(len(colleges)), [c[:20] + '...' if len(c) > 20 else c for c in colleges], rotation=45)
            plt.grid(True, alpha=0.3)
            
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close()
            charts['satisfaction_by_college'] = buffer
        
        # Barriers Analysis Chart
        barriers_data = []
        for resp in data['responses']:
            barriers = resp.get('barriers', {})
            barriers_data.append({
                'Financial': barriers.get('financialBarrier', 0),
                'Technical': barriers.get('technicalBarrier', 0),
                'Training': barriers.get('trainingBarrier', 0),
                'Policy': barriers.get('policyBarrier', 0)
            })
        
        if barriers_data:
            fig, ax = plt.subplots(figsize=(10, 6))
            barriers_df = pd.DataFrame(barriers_data)
            barriers_mean = barriers_df.mean()
            
            plt.bar(barriers_mean.index, barriers_mean.values, color=['red', 'orange', 'yellow', 'green'])
            plt.title('Average Barrier Scores')
            plt.xlabel('Barrier Type')
            plt.ylabel('Average Score')
            plt.grid(True, alpha=0.3)
            
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close()
            charts['barriers_analysis'] = buffer
        
        return charts
    
    def _add_summary_to_pdf(self, story, summary: Dict[str, Any], styles):
        """Add summary section to PDF"""
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Responses', str(summary.get('totalResponses', 0))],
            ['Average Infrastructure Score', f"{summary.get('averageInfrastructureScore', 0):.2f}"],
            ['Average Satisfaction', f"{summary.get('averageSatisfaction', 0):.2f}"],
            ['Critical Barriers Count', str(summary.get('criticalBarriersCount', 0))],
            ['Colleges Covered', str(summary.get('collegesCount', 0))],
        ]
        
        table = Table(summary_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 12))


def get_report_generator() -> 'ReportGenerator':
    """Factory to obtain a ReportGenerator instance.

    Callers should use get_report_generator() instead of importing a module
    level instance. This avoids running plotting imports at import time which
    could prevent the app from starting if plotting libs are missing.
    """
    return ReportGenerator()