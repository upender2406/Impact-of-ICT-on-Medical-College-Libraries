"""
Import survey data from the Final Survey_Responses0.xlsx file into the database.

This script:
1. Clears existing survey_responses (preserving users, notifications, audit_logs)
2. Imports all rows from the xlsx file into the survey_responses table
3. Maps xlsx columns to the SurveyResponse SQLAlchemy model
"""
import sys
from pathlib import Path

# Add app directory to path
sys.path.insert(0, str(Path(__file__).parent))

from datetime import datetime
import openpyxl
from app.database import engine, SessionLocal, Base
from app.models.db_models import SurveyResponse, SubmissionStatus, AutomationSystem, RespondentType

XLSX_PATH = Path(__file__).parent / "database" / "Final Survey_Responses0.xlsx"

# Mapping from xlsx automation_system strings to the enum
AUTOMATION_MAP = {
    "NONE": AutomationSystem.NONE,
    "None": AutomationSystem.NONE,
    None: AutomationSystem.NONE,
    "KOHA": AutomationSystem.KOHA,
    "SOUL": AutomationSystem.SOUL,
    "OTHER": AutomationSystem.OTHER,
    "Other": AutomationSystem.OTHER,
}

# Mapping from xlsx respondent_type strings to the enum
RESPONDENT_MAP = {
    "STUDENT": RespondentType.STUDENT,
    "Student": RespondentType.STUDENT,
    "FACULTY": RespondentType.FACULTY,
    "Faculty": RespondentType.FACULTY,
    "RESEARCHER": RespondentType.RESEARCHER,
    "Researcher": RespondentType.RESEARCHER,
    "LIBRARY_STAFF": RespondentType.LIBRARY_STAFF,
    "Library_Staff": RespondentType.LIBRARY_STAFF,
}

# Mapping from xlsx status strings to the enum
STATUS_MAP = {
    "PENDING": SubmissionStatus.PENDING,
    "pending": SubmissionStatus.PENDING,
    "APPROVED": SubmissionStatus.APPROVED,
    "approved": SubmissionStatus.APPROVED,
    "REJECTED": SubmissionStatus.REJECTED,
    "rejected": SubmissionStatus.REJECTED,
    "REVISION_REQUESTED": SubmissionStatus.REVISION_REQUESTED,
    "revision_requested": SubmissionStatus.REVISION_REQUESTED,
}


def parse_datetime(val):
    """Parse a datetime value from the xlsx cell."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def import_xlsx():
    """Import xlsx data into the database."""
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx file not found at {XLSX_PATH}")
        sys.exit(1)

    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Read headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    print(f"Headers ({len(headers)}): {headers[:5]}...")

    db = SessionLocal()
    try:
        # Clear existing survey responses
        existing_count = db.query(SurveyResponse).count()
        print(f"\nClearing {existing_count} existing survey responses...")
        db.query(SurveyResponse).delete()
        db.commit()

        # Import rows
        imported = 0
        errors = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))

            try:
                response = SurveyResponse(
                    # Skip Sl.No. – use auto-increment id
                    user_id=row_data.get("user_id"),
                    status=STATUS_MAP.get(row_data.get("status"), SubmissionStatus.PENDING),
                    reviewed_by=row_data.get("reviewed_by"),
                    reviewed_at=parse_datetime(row_data.get("reviewed_at")),
                    review_notes=row_data.get("review_notes"),
                    version=int(row_data.get("version", 1) or 1),
                    college=row_data.get("college", "Unknown"),
                    college_tier=row_data.get("college_tier"),
                    respondent_type=RESPONDENT_MAP.get(row_data.get("respondent_type"), RespondentType.STUDENT),
                    respondent_name=row_data.get("respondent_name"),
                    respondent_position=row_data.get("respondent_position"),
                    respondent_email=row_data.get("respondent_email"),
                    hardware_quality=float(row_data.get("hardware_quality", 0) or 0),
                    software_availability=float(row_data.get("software_availability", 0) or 0),
                    internet_speed=float(row_data.get("internet_speed", 0) or 0),
                    digital_collection=float(row_data.get("digital_collection", 0) or 0),
                    automation_system=AUTOMATION_MAP.get(row_data.get("automation_system"), AutomationSystem.NONE),
                    infrastructure_score=float(row_data.get("infrastructure_score", 0) or 0),
                    overall_satisfaction=float(row_data.get("overall_satisfaction", 0) or 0),
                    service_efficiency=float(row_data.get("service_efficiency", 0) or 0),
                    staff_helpfulness=float(row_data.get("staff_helpfulness", 0) or 0),
                    financial_barrier=float(row_data.get("financial_barrier", 0) or 0),
                    technical_barrier=float(row_data.get("technical_barrier", 0) or 0),
                    training_barrier=float(row_data.get("training_barrier", 0) or 0),
                    policy_barrier=float(row_data.get("policy_barrier", 0) or 0),
                    barrier_score=float(row_data.get("barrier_score", 0) or 0),
                    weekly_visits=int(row_data.get("weekly_visits", 0) or 0),
                    ict_training_received=bool(int(row_data.get("ict_training_received", 0) or 0)),
                    awareness_level=int(row_data.get("awareness_level", 3) or 3),
                    remote_access_available=bool(int(row_data.get("remote_access_available", 0) or 0)),
                    digital_resource_usage=row_data.get("digital_resource_usage"),
                    pandemic_adaptation=row_data.get("pandemic_adaptation"),
                    comments=row_data.get("comments"),
                    anomaly_score=float(row_data["anomaly_score"]) if row_data.get("anomaly_score") is not None else None,
                    quality_score=float(row_data["quality_score"]) if row_data.get("quality_score") is not None else None,
                    submitted_at=parse_datetime(row_data.get("submitted_at")) or datetime.utcnow(),
                    updated_at=parse_datetime(row_data.get("updated_at")),
                )
                db.add(response)
                imported += 1

                # Commit in batches of 500
                if imported % 500 == 0:
                    db.commit()
                    print(f"  ...imported {imported} rows")

            except Exception as e:
                errors += 1
                row_id = row_data.get("id", "?")
                print(f"  ERROR on row id={row_id}: {e}")
                db.rollback()

        # Final commit
        db.commit()
        wb.close()

        # Verify
        final_count = db.query(SurveyResponse).count()
        print(f"\n{'=' * 60}")
        print(f"IMPORT COMPLETE")
        print(f"  Imported: {imported}")
        print(f"  Errors:   {errors}")
        print(f"  Total survey_responses in DB: {final_count}")
        print(f"{'=' * 60}")

    except Exception as e:
        db.rollback()
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    import_xlsx()
